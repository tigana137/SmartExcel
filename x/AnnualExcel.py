import re
from openpyxl import load_workbook
from x.functions import CustomError, check_if_sid_need_to_be_replaced
from x.models import AdminEcoledata


column_starting_point = 'B'
row_starting_point = 5



def check_if_sid_valid(sid:str,row:int):
    r'''check if the school id is digit with a length of 6'''

    pattern = re.compile(r'^\d{6}$')
    is_valid_sid = bool(pattern.match(sid))

    if not is_valid_sid :
        raise CustomError(f'''fama id t3 ecole mehouch  digit walla length te3ou mouch 6 f row {row}
            l valeur ta3 l cell is : "{sid}"
            ''')



def check_duplicates(sid,sids_array,ministre_school_name):
    r'''check if id is already treated b4 in the excel file'''
    if sid in sids_array:
        confirmation_msg =input(f'''l id {sid} t3 l'ecole {ministre_school_name} deja 3mltlha l traitement 9bal donc imma hedhi l id
        t3ha 8alt walla lo5ra press 'y' to break : ''')
        if confirmation_msg=='y':
            raise CustomError('')



def check_if_sid_in_db(sid,sids,ministre_school_name:str,):

    if int(sid) not in sids :
        confirmation_msg =input(f"l id {sid} t3 l'ecole {ministre_school_name} mouch mawjoud fil db press 'y' to break : ")
        if confirmation_msg=='y':
            raise CustomError('')
    return

def check_if_cell_null(nbr_elvs,nbr_classes,niveau,row):
    # if not nbr_elvs :
    #     raise CustomError(f'''nbr elvs null f niveau {niveau} f row {row}''')
    return

def annualexcel(dre_id:int):
    sids = AdminEcoledata.objects.filter(dre_id=dre_id).values_list('sid',flat=True)
    wb = load_workbook("annualexcel.xlsx",data_only=True)
    ws = wb.active

    row = row_starting_point
    
    ecoles = []
    sids_array = []
    
    while ws['D'+str(row)].value :
        sid = str(ws['D'+str(row)].value)
        sid = check_if_sid_need_to_be_replaced(str(sid),public_school=True)

        check_if_sid_valid(sid,row)

        ministre_school_name = ws['E'+str(row)].value
        
        check_if_sid_in_db(sid,sids,ministre_school_name)
        check_duplicates(sid,sids_array,ministre_school_name)


        premier_elvs = ws['K'+str(row)].value
        premier_classes = ws['L'+str(row)].value
        check_if_cell_null(premier_elvs,premier_classes,niveau=1,row=row)

        deuxieme_elvs = ws['P'+str(row)].value
        deuxieme_classes = ws['Q'+str(row)].value
        check_if_cell_null(deuxieme_elvs,deuxieme_classes,niveau=2,row=row)

        troisieme_elvs = ws['U'+str(row)].value
        troisieme_classes = ws['V'+str(row)].value
        check_if_cell_null(troisieme_elvs,troisieme_classes,niveau=3,row=row)

        quaterieme_elvs = ws['Z'+str(row)].value
        quaterieme_classes = ws['AA'+str(row)].value
        check_if_cell_null(quaterieme_elvs,quaterieme_classes,niveau=4,row=row)

        cinquieme_elvs = ws['AE'+str(row)].value
        cinquieme_classes = ws['AF'+str(row)].value
        check_if_cell_null(cinquieme_elvs,cinquieme_classes,niveau=5,row=row)

        sixieme_elvs = ws['AJ'+str(row)].value
        sixieme_classes = ws['AK'+str(row)].value
        check_if_cell_null(sixieme_elvs,sixieme_classes,niveau=6,row=row)
      
        if row != 144:
            AdminEcoledata.objects.get(sid=sid).update_levelstat((premier_elvs,premier_classes),(deuxieme_elvs,deuxieme_classes),(troisieme_elvs,troisieme_classes),(quaterieme_elvs,quaterieme_classes),(cinquieme_elvs,cinquieme_classes),(sixieme_elvs,sixieme_classes))        
        
        ecole = AdminEcoledata(
            sid=sid,
            ministre_school_name=ministre_school_name 
        )
        ecoles.append(ecole)
        sids_array.append(sid)
        row+=1

    # AdminEcoledata.objects.bulk_update(ecoles,fields=['ministre_school_name'])
    
    return



