import requests
from rest_framework.decorators import api_view,permission_classes
import base64
from django.http import JsonResponse
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from excel.functions import adjust_levelstat
from excel.models import excelsheets, excelsheets_brillant
from users.functions import verify_jwt
from users.models import UserProfile
from x.AnnualExcel import annualexcel
from x.UpdateStudents import UpdateStudents
from x.UpdateSchools import UpdateSchools
from x.Update_1st_grade_students import Update_1st_grade_students
from x.Update_kindergarten_students import Update_kindergarten_students
from x.excelPhoneNumberEmail import excelPhoneNumberEmail
from x.reset_dre_database import reset_dre_database
from x.UpdatesPrincipals import update_principals
from x.exportModels import exportAdminEcoledata, exportAdminElvs, exportDel1, exportDre, exportElvsprep, exportExcelSheets, exportbrillantExcelSheets, exportlevelstat
from x.functions import CustomError, get_sids_to_replace
from x.importModels import importAdminEcoledata, importAdminElvs, importBrillantExcelSheets, importDel1, importDre, importElvsprep, importExcelSheets, importlevelstat, importlevelstat2
from x.models import AdminEcoledata, AdminElvs, Del1, DirtyNames, Dre, Elvsprep, Tuniselvs, levelstat
from openpyxl import load_workbook




request2 = requests.session()

import re

@api_view(['GET'])
@permission_classes([AllowAny])
def testSignal(request):
    # "http://localhost:80/api/x/testSignal/"
    return Response(True)  



@api_view(['GET'])
def testAuth(request):
    "http://localhost:80/api/x/testAuth/"
    return Response(True) 


 
@api_view(['GET'])
@permission_classes([AllowAny])
def GetCapatcha(request):
    "http://localhost:80/api/x/GetCapatcha/"


    url = "https://suivisms.cnte.tn/"
    request2.get(url=url)

    try:
        url_img = "https://suivisms.cnte.tn/inclure/img.php"
        img = request2.get(url=url_img)
    except Exception as e:
        print(e)
        raise CustomError("fama mochkla k jit te5ou fil taswira")

    # Assuming you have the binary image data in 'image_data'
    image_data_base64 = base64.b64encode(img.content).decode('utf-8')

    # Create a JSON response with the base64-encoded image
    response_data = {'image_data': image_data_base64}

    with open('capatcha_img.jpg',"wb") as w :
        w.write(img.content)

    return JsonResponse(response_data)

from bs4 import BeautifulSoup as bs

@api_view(['GET'])
@permission_classes([AllowAny])
def VerifyCapatcha(request, code):
    "http://localhost:80/api/x/VerifyCapatcha/"

    url = "https://suivisms.cnte.tn/"
    payload = {"login": "user8420",
               "pwd": "78b9adE48U",
               "secure": code,
               "auth": "",
               }

    response = request2.post(url=url, data=payload)

    # response.headers
    if not ("https://suivisms.cnte.tn/" in response.url):
        print(response.url)
        return Response(False)
    try:
        0
        id = '015666049361'

        # tmchi l page l t7wil 
        request2.get("https://suivisms.cnte.tn/ministere/index.php?op=inscprim&act=find_mvt",data={'op':'inscprim','act':'find_mvt'})
        
        # tcharchi 3al telmidh
        response = request2.post("https://suivisms.cnte.tn/ministere/index.php?op=inscprim&act=mvt",data={'idenelev':id,'btenv':'بحث'})
        soup = bs(response.content.decode(encoding='utf-8', errors='ignore'), 'html.parser')

        prev_sid = soup.find('input',{'type':'hidden','name':'codeetab'})
        
        if not prev_sid:
            raise CustomError('me3ndouch previous school id f site 8riba')

        prev_sid= prev_sid['value']


        if not soup.find('select',{'name':'code_etab2'}):
            raise CustomError('elv not found')
       
        sids_to_replace =get_sids_to_replace()
        next_sid = 842823

        if str(next_sid) in sids_to_replace.values():
            next_sid = sids_to_replace.inv[str(next_sid)]

        # to transfer a elv
        # next_sid = 842401
        print(f'id ={id} , prev_sid = {prev_sid} , next_sid : {next_sid}')
        # return
        payload = {'idenuniq':id,'codeetab':prev_sid,'code_etab2':next_sid , 'btenv':'الموافقة على نقلة التلميذ '}
        response =request2.post("https://suivisms.cnte.tn/ministere/index.php?op=inscprim&act=do_mvt",data=payload)
        soup = bs(response.content.decode(encoding='utf-8', errors='ignore'), 'html.parser')
        if "لقد تمت عملية النقلة بنجاح" in str(soup) : 
            print('cbn')
        else:
            print(str(soup)) 
            print('nope')
        # dre = reset_dre_database(request2) 
        # UpdateSchools(request2,dre)
        # UpdateStudents(request2,dre)
        # Update_1st_grade_students(request2,dre)
        # Update_kindergarten_students(request2,dre)
        # annualexcel(dre_id=dre.id)
        # print("\n✓✓✓ operation completed succesfully\n")
    finally:
        request2.close()

    return Response(True)



@api_view(['GET'])
def getMoudirins(request):
    update_principals()

    return Response(True)


@api_view(['GET'])
@permission_classes([AllowAny])
def exportDB(request):
    "http://localhost:80/api/x/exportDB"
    # exportDre()
    # exportDel1()
    # exportlevelstat()
    exportAdminEcoledata()  
    # exportAdminElvs()
    # exportElvsprep()
    # exportExcelSheets()
    # exportbrillantExcelSheets()
    return Response(True)


@api_view(['GET'])
@permission_classes([AllowAny])
def importDB(request):
    "http://localhost:80/api/x/importDB"
    # # AdminElvs.objects.all().delete()
    # # Elvsprep.objects.all().delete()
    # importDre()
    # importDel1()
    # importlevelstat()
    importAdminEcoledata()
    # importAdminElvs()
    # importElvsprep()
    # excelsheets.objects.all().delete()
    # importExcelSheets()
    # excelsheets_brillant.objects.all().delete()
    # importBrillantExcelSheets()
    return Response(True) 


@api_view(['GET'])
def updateLevelStat(request):
    r'''this one updates only the inital number of elves in 30 august and number of classes in case the annuel excel
    comes out late and they are transferring students already'''
    "http://localhost:80/api/x/updateLevelStat/"
    # importlevelstat2()
    return Response(True)





@api_view(['GET'])
def updateSchoolPhoneNumbers(request):
    r'''this one updates school's phone numbers and email from an excel file
    the file name should be : PhonesAndEmails.xlsx'''
    "http://localhost:80/api/x/updateSchoolPhoneNumbers/"
    # excelPhoneNumberEmail(dre_id=84)
    return Response(True)



@api_view(['GET'])
@permission_classes([AllowAny])
def updateExcelSheets_brillant(request):
    "http://localhost:80/api/x/updateExcelSheets_brillant/" 

    # return Response(True)
    excel_name = ".xlsx"         # !!!!
    date_downloaded = "2024-09-16"    # !!!! YYYY-MM-DD
    
    starrin_rows_each_level = {     # !!!!
        1:8,
        2:8,
        3:8,
        4:8,
        5:8,
        6:8,
    }
    starrin_charr_each_level = {    # !!!!
        1:'C',
        2:'C',
        3:'C',
        4:'C',
        5:'C',
        6:'C',
    }


    wb = load_workbook(excel_name,data_only=True)

    array_excelsheets_brillant= []

    for elv_level in range(1,7):

        ws = wb.worksheets[elv_level-1]
        row_starting_point= starrin_rows_each_level[elv_level]
        charr =starrin_charr_each_level[elv_level]

        row = row_starting_point 

        while ws[charr+str(row)].value or ws[charr+str(row+1)].value or ws[charr+str(row+2)].value or ws[charr+str(row+3)].value :

            Del1 = str(ws[charr+str(row)].value)
            
            next_char = chr(ord(charr) + 1)
            nom_prenom = str(ws[next_char+str(row)].value)

            next_char = chr(ord(next_char) + 1)
            uid = str(ws[next_char+str(row)].value)

            next_char = chr(ord(next_char) + 1)
            prev_ecole = str(ws[next_char+str(row)].value)
        
            next_char = chr(ord(next_char) + 1)
            next_ecole = str(ws[next_char+str(row)].value)

            # next_char = chr(ord(next_char) + 1)
            # reason = str(ws[next_char+str(row)].value)

            next_char = chr(ord(next_char) + 1)
            decision = str(ws[next_char+str(row)].value)

             
            sheet = excelsheets_brillant(
                uid=uid,
                Del1=Del1,
                nom_prenom=nom_prenom,
                prev_ecole=prev_ecole,
                next_ecole=next_ecole,
                level=elv_level,
                reason="  ",
                decision=decision,
                date_downloaded=date_downloaded,
                dre_id=84
            )
            row+=1
            array_excelsheets_brillant.append(sheet)
    # print(f' uid = --{uid}--\n Del1 = {Del1} \n nom_prenom = {nom_prenom} \n prev_ecole = {prev_ecole} \n next_ecole = {next_ecole} \n level = {elv_level}')
    print(len(array_excelsheets_brillant))
    # excelsheets_brillant.objects.bulk_create(array_excelsheets_brillant,)

    return Response(True)




 

@api_view(['GET'])
@permission_classes([AllowAny])
def updateExcelSheets_brillant_singlesheet(request):
    "http://localhost:80/api/x/updateExcelSheets_brillant_singlesheet/" 

    # return Response(True)
    date = "" 
    excel_name = date+".xlsx"         # !!!!
    date_downloaded = "2024-"+date    # !!!! YYYY-MM-DD

    wb = load_workbook(excel_name,data_only=True)

    array_excelsheets_brillant= []

    ws = wb.worksheets[0]
    row_starting_point= 5
    charr = "A"



    row = row_starting_point

    while ws[charr+str(row)].value or ws[charr+str(row+1)].value or ws[charr+str(row+2)].value or ws[charr+str(row+3)].value :

        level = str(ws[charr+str(row)].value)

        next_char = chr(ord(charr) + 1)
        next_char = chr(ord(next_char) + 1)
        Del1 = str(ws[next_char+str(row)].value) 
        
        next_char = chr(ord(next_char) + 1)
        nom_prenom = str(ws[next_char+str(row)].value)

        next_char = chr(ord(next_char) + 1)
        uid = str(ws[next_char+str(row)].value)

        next_char = chr(ord(next_char) + 1)
        prev_ecole = str(ws[next_char+str(row)].value)
    
        next_char = chr(ord(next_char) + 1)
        next_ecole = str(ws[next_char+str(row)].value)

        # next_char = chr(ord(next_char) + 1)
        # reason = str(ws[next_char+str(row)].value)

        next_char = chr(ord(next_char) + 1)
        decision = str(ws[next_char+str(row)].value)

            
        sheet = excelsheets_brillant(
            uid=uid,
            Del1=Del1,
            nom_prenom=nom_prenom,
            prev_ecole=prev_ecole,
            next_ecole=next_ecole,
            level=level if level else 3,
            reason="  ",
            decision=decision,
            date_downloaded=date_downloaded,
            dre_id=84
        )
        row+=1
        array_excelsheets_brillant.append(sheet)
    print(f' uid = --{uid}--\n Del1 = {Del1} \n nom_prenom = {nom_prenom} \n prev_ecole = {prev_ecole} \n next_ecole = {next_ecole} \n level = {level}')
    print(f'prev count of excels = {excelsheets_brillant.objects.all().count()}')
    print(len(array_excelsheets_brillant))
    # excelsheets_brillant.objects.bulk_create(array_excelsheets_brillant,)
    # print(f'new count of excels = {excelsheets_brillant.objects.all().count()}')


    return Response(True) 



def consists_of_12_digits(s: str) -> bool:
    return True
    # Check if the string is exactly 12 characters long and contains only digits
    return len(s) == 12 and s.isdigit() or len(s) == 11 and s.isdigit() 



def alterstr(uid,):
    
    if type(uid) == str : 
        uid = uid.strip()

    if  type(uid) == str and uid.endswith("'0"):
        uid = uid[:-2]

    if type(uid) == str :
        uid = uid.translate(str.maketrans("", "", ", ' *"))

    if type(uid) == str and len(uid) ==12 and  uid.isdigit() and uid.startswith('01'):
        return uid
    
    if  type(uid) == str and len(uid) ==11 and  uid.isdigit() and uid.startswith('1'):
        return uid
    
    

 
    return 0

 

def add_new_resttt_excel_row(del1,nom_prenom,uid,prev_ecole,next_ecole,reason,decision,level,problem):


    print(problem)

    workbook = load_workbook("rest_new.xlsx")
    sheetname = workbook.sheetnames[0]
    sheet = workbook[sheetname]
    charr ='A'
    row = dicc[int(level)]
    
    sheet[charr+str(row)] = del1
            
    next_char = chr(ord(charr) + 1)
    sheet[next_char+str(row)] = nom_prenom
 
    next_char = chr(ord(next_char) + 1)
    sheet[next_char+str(row)] = uid

    next_char = chr(ord(next_char) + 1)
    sheet[next_char+str(row)] = level

    next_char = chr(ord(next_char) + 1)
    sheet[next_char+str(row)] = prev_ecole

    next_char = chr(ord(next_char) + 1)
    sheet[next_char+str(row)] = next_ecole

    next_char = chr(ord(next_char) + 1)
    sheet[next_char+str(row)] = reason

    next_char = chr(ord(next_char) + 1)
    sheet[next_char+str(row)] = decision

    next_char = chr(ord(next_char) + 1)
    sheet[next_char+str(row)] = problem

    dicc[int(level)]+=1

    workbook.save("rest_new.xlsx")

 

@api_view(['GET'])
@permission_classes([AllowAny])
def transfer_rest(request):
    "http://localhost:80/api/x/transfer_rest/" 
  
    excels = excelsheets_brillant.objects.all()

    x=0

    decisions_dic = ["مع الموافقة","مع المولفقة","مع النوافقة","موافقة المدير",]

    ecoles = AdminEcoledata.objects.exclude(sid__startswith = 8498)
    ecoles_array = {}
    for ecole in ecoles :
        ecoles_array[ecole.ministre_school_name.replace(' ','')] = ecole.sid
    

    for excelsheet in excels:
        uid = excelsheet.uid
        uid = alterstr(uid)

        if uid == 0 :
            x+=1
            continue


        nextEcole = excelsheet.next_ecole if excelsheet.decision in decisions_dic else excelsheet.decision
        # print(nextEcole)

        ecole = AdminEcoledata.objects.filter(ministre_school_name=nextEcole).first()

        if not ecole :
            nextEcole = str(nextEcole).replace(' ','')
            if nextEcole not in ecoles_array:
                x+=1
                # print(nextEcole)

    print(x)
    return Response(True)

