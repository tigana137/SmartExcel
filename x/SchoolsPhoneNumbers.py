from openpyxl import load_workbook
from x.functions import CustomError
from x.models import AdminEcoledata


def SchoolsPhoneNumbers(dre_id:int):
    r'''extract from an excel school's email and phone number 
    excel's name should be ...

    A column : sid
    B column : phone number
    C column : email

    '''
    sids = AdminEcoledata.objects.filter(dre_id=dre_id).values_list('sid',flat=True)
    wb = load_workbook("xx.xlsx")
    ws = wb.active

    starting_row =2 # l awil star fih sids and info


    data = {}
    row = starting_row

    while ws['A'+str(starting_row)].value:
        
        sid = ws['A'+str(starting_row)].value
        numero = ws['B'+str(starting_row)].value
        email = ws['C'+str(starting_row)].value
        
        if sid in data.keys():
            raise CustomError('duplicate sid : ' +str(sid))

        data[sid]= {
            'numero':numero,
            'email':email,
        }

        row+=1

    