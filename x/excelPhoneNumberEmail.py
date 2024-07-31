import re
from openpyxl import load_workbook
from x.AnnualExcel import check_duplicates, check_if_sid_in_db, check_if_sid_valid
from x.functions import CustomError, check_if_sid_need_to_be_replaced
from x.models import AdminEcoledata





def is_valid_phone_number(phone_number,row):
    r'''returns None is the phone number is empty or the number if it s valid or a confirmation msg if it s not'''
    if phone_number==None:
        return None
    
    try:
        phone_number = int(phone_number)
    except:
        confirmation_msg=input(f'''l cell t3 phone number f row {row} mehyech kolh ar9am if you wanna just skip press 'y' : ''')
        if confirmation_msg=='y':
            return None
        raise CustomError('')

    pattern = re.compile(r'^\d{8}$')

    is_valid = bool(pattern.match(str(phone_number)))
    if is_valid:
        return phone_number
    
    confirmation_msg =input(f''' l ecole f row {row} l pattern t3 l phone number 8alta press 'y' if you want to continue and ignore it : ''')
    if confirmation_msg=='y':
        return None
    
    raise CustomError('')       
     

def is_valid_email(email,row):
    # Define a pattern for a valid email address
    pattern = re.compile(
        r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    )
    if email==None:
        return None
    
    pattern = re.compile(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$')

    is_valid = bool(pattern.match(email))
    if is_valid:
        return email
    
    confirmation_msg =input(f''' l ecole f row {row} l pattern t3 l email 8alta press 'y' if you want to continue and ignore it : ''')
    if confirmation_msg=='y':
            return None  
    
    raise CustomError('')     




row_starting_point = 2

def excelPhoneNumberEmail(dre_id:int):
    r'''extract from an excel school's phone numbers and email
        the excel structure is first row is table head 
        A1 : sid , A2 : principal's name , A3 : Principal's phone1 , A4 : Principal's phone2 , A5 : School's email'''
    print('\n** Updating Schools phone numbers and emails ')


    sids = AdminEcoledata.objects.filter(dre_id=dre_id).values_list('sid',flat=True)
    wb = load_workbook("PhonesAndEmails.xlsx")
    ws = wb.active

    row = row_starting_point

    ecoles=[]
    sids_array = []

    while ws['A'+str(row)].value:

        sid = str(int(ws['A'+str(row)].value))
        sid = check_if_sid_need_to_be_replaced(sid,public_school=True)
        check_if_sid_valid(sid,row)
        check_if_sid_in_db(sid,sids,None) # if you asking why none 5atr l function 3mltha melloul l annuel excel 3ala eses ken fama erruer t5arrjlk ism l madrsa welli inti me3ndkch fil excel hedha
        check_duplicates(sid,sids_array,None)

        principal = str(ws['B'+str(row)].value)
        phone1 = ws['C'+str(row)].value
        phone2 = ws['D'+str(row)].value
        email = str(ws['E'+str(row)].value)

        phone1 = is_valid_phone_number(phone1,row)
        phone2 = is_valid_phone_number(phone2,row)
        # email = is_valid_email(email,row)

        ecole= AdminEcoledata(
             sid=sid,
             phone1=phone1,
             phone2=phone2,
             email=email,
        )
        if principal:
             ecole.principal = principal
        
        ecoles.append(ecole)
        sids_array.append(sid)
        row+=1

    AdminEcoledata.objects.bulk_update(ecoles,fields=["principal","phone1","phone2","email"])

    print("✓ Schools phone numbers and emails updated succesfully")
    






